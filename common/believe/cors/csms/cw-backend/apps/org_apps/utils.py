import csv, traceback, hashlib, json, requests, xlrd, xlsxwriter, uuid, os, re, tempfile
from datetime import datetime

from io import StringIO, BytesIO
from openpyxl import load_workbook

from django.db import models
from django.conf import settings
from django.urls import reverse
from django.utils.translation import gettext as _
from django.utils import timezone
from django.core.files.uploadedfile import SimpleUploadedFile
from requests.auth import HTTPBasicAuth
from rest_framework import status
from rest_framework.response import Response
from xlrd import XL_CELL_BOOLEAN, XL_CELL_DATE, XL_CELL_NUMBER, xldate
from dateutil import *
from dateutil.tz import *
from dateutil.rrule import MONTHLY, rrule

import process_engine
from django.dispatch import receiver
from ezedox.settings import AWS_STORAGE_BUCKET_NAME, REPORT_BATCH_SIZE, HIGH_PRIORITY_TASK, BASE_ORG_DOMAIN_URL, CANDIDATE_DOMAIN_URL, BACKEND_DOMAIN_URL, DEFAULT_SCHEME, PLATFORM_BASE_URL, PLATFORM_INTERNAL_TOKEN
from ezedox.celery import app
from ezedox.settings import PROCESS_ENGINE_PASSWORD, PROCESS_ENGINE_USER, DEFAULT_SCHEME, SSL_VERIFICATION
from apps.org_apps.models import OrganisationWorkflow, ProcessVarSync
from apps.org_form.models import OrganisationForm, Transaction, OrganisationFile, generate_path
from apps.org_form.utils import get_key_type
from apps.org_users.utils import get_tenant
from apps.org_group.models import OrganisationGroup
from apps.org_users.serializers import OrgUserSerializer
from apps.organisations.models import Organisation
from apps.org_entity.models import OrganisationEntityMasterData
from apps.org_import.models import EntityImport
from apps.proxy_bpm.utils import replace_newlines
from apps.proxy_bpm.flowable import get_process_by_process_ids, get_var_by_process_id
from apps.organisations.serializers import OrganisationDataInSerializer
from apps.org_bff.services.employee_mgmt import EmployeeMgmtService
from apps.org_bff.services.vendor_mgmt import VendorMgmtService
from apps.org_bff.services.customer_mgmt import CustomerMgmtService
from utils import storage_utils
from utils.loggerwrapper import Logger, getMessage, getLogMessage
from utils.process_engine_proxy import call
from .internal_errors import org_apps_errors


logger = Logger('__name__')

COMPARISION = {
    "EQUALS" : "equals",
    "NOT_EQUALS" : "notEquals",
    "NOT_EQUALS_IGNORE_CASE" : "notEqualsIgnoreCase",
    "EQUALS_IGNORE_CASE" : "equalsIgnoreCase",
    "GREATER_THAN" : "greaterThan",
    "GREATER_THAN_OR_EQUALS" : "greaterThanOrEquals",
    "LESS_THAN": "lessThan",
    "LESS_THAN_OR_EQUALS" : "lessThanOrEquals",
    "LIKE" : "like",
    "LIKE_IGNORE_CASE" : "likeIgnoreCase"
}

def months(start_month, start_year, end_month, end_year):
    start = datetime(start_year, start_month, 1)
    end = datetime(end_year, end_month, 1)
    return [(d.month, d.year) for d in rrule(MONTHLY, dtstart=start, until=end)]

def get_value(json_data, json_key):
    try:
        value = json_data
        for item in json_key.split('.'):
            value = value[item]
        return value
    except Exception as error:
        return None


def get_process_name(process_name, tenant, workflow, var_data):
    name = ""
    name_list = process_name.split(',')
    for name_list_item in name_list:
        if name_list_item in var_data:
            name = name +  var_data.get(name_list_item) + "_"
        elif name_list_item == "workflow_id":
            name = name + workflow + "_"
        elif name_list_item == "tenant_id":
            name = name + tenant + "_"
        else:
            slash_split = name_list_item.split('/')
            for slash_split_item in slash_split:
                if get_value(var_data, slash_split_item):
                    name = name +  get_value(var_data, slash_split_item) + "_"
                    break
    if name == "":
        return None
    name = name[:-1]
    return name


def get_uploaded_file(f):
    file_name = ''
    if f.name.endswith('.csv'):
        suffix = ".csv"
    elif f.name.endswith('.xlsx'):
        suffix = ".xlsx"
    with tempfile.NamedTemporaryFile(
            suffix=suffix, delete=False) as dest_file:
        file_name = dest_file.name
        for chunk in f.chunks():
            dest_file.write(chunk)
    return file_name

def xls_to_csv(filepath):
    """
    Convert excel to csv
    """
    excel = xlrd.open_workbook(filepath)
    sheet_1 = excel.sheet_by_index(0)
    datas = []

    def _parse_cell(cell_content, cell_type):
        """
        Parse the cell components type to native type
        """
        if cell_type == XL_CELL_DATE:
            cell_content = xldate.xldate_as_datetime(cell_content, excel.datemode)
            cell_content = cell_content.strftime("%d %b %Y")
        elif cell_type == XL_CELL_BOOLEAN:
            cell_content = bool(cell_content)
        elif cell_type == XL_CELL_NUMBER:
            cell_content = int(cell_content)
        return cell_content

    for row in range(sheet_1.nrows):
        rows = []
        for index, (value, type) in enumerate(zip(sheet_1.row_values(row), sheet_1.row_types(row))):
            rows.append(_parse_cell(value, type))
        datas.append(rows)
    with tempfile.NamedTemporaryFile(mode='w', encoding="utf8", delete=False, suffix='.csv', dir=settings.TEMP_DIR) as csv_file:
        csv_writer = csv.writer(csv_file)
        csv_writer.writerows(datas)
        return csv_file.name

def xls_data(filepath):
    excel = xlrd.open_workbook(filepath)
    sheet_1 = excel.sheet_by_index(0)
    datas = []

    def _parse_cell(cell_content, cell_type):
        """
        Parse the cell components type to native type
        """
        if cell_type == XL_CELL_DATE:
            cell_content = xldate.xldate_as_datetime(cell_content, excel.datemode)
            cell_content = cell_content.strftime("%d %b %Y")
        elif cell_type == XL_CELL_BOOLEAN:
            cell_content = bool(cell_content)
        elif cell_type == XL_CELL_NUMBER:
            cell_content = int(cell_content)
        return cell_content

    for row in range(sheet_1.nrows):
        rows = []
        for index, (value, type) in enumerate(zip(sheet_1.row_values(row), sheet_1.row_types(row))):
            rows.append(_parse_cell(value, type))
        datas.append(rows)

    return datas


def process_row(data_type, key, row):
    new_row = []
    for index, value in enumerate(row):
        if data_type[index] == 'string':
            new_row.append(str(value).strip())
        elif data_type[index] == 'json':
            try:
                new_row.append(json.loads(value))
            except Exception as e:
                new_row.append(value)
        elif data_type[index] == 'boolean':
            new_row.append(value)
        else:
            new_row.append(value)
    try:
        index = key.index('entity_email')
        new_row[index] = new_row[index].lower()
    except Exception:
        pass
    return new_row

@app.task(bind=True, name="bulk_initiate_process")
def bulk_task_create(self, request_body, entity_import_id, tenant_id):
    import_obj = EntityImport.objects.get(id=entity_import_id, tenant=tenant_id)
    tenant = Organisation.objects.get(id=tenant_id)
    if import_obj:
        try:
            data = []
            blob_name = import_obj.file.name  
            blob_content = storage_utils.read_blob_content(blob_name)
            if request_body['file_type'] == '.csv':
                blob_str = StringIO(blob_content.decode('utf-8'))
                csv_data_rows = csv.reader(blob_str)
                for csv_data_row in csv_data_rows:
                    data.append(csv_data_row)
            elif request_body['file_type'] == '.xlsx':
                blob_stream = BytesIO(blob_content)
                workbook = load_workbook(blob_stream)
                sheet = workbook.active
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row))
            elif request_body['file_type'] == '.xls':
                blob_stream = BytesIO(blob_content)
                workbook = xlrd.open_workbook(file_contents=blob_stream.read())
                sheet = workbook.sheet_by_index(0)
                for row_idx in range(sheet.nrows):
                    row = sheet.row_values(row_idx)
                    data.append(row)
            total_count = 0
            success_count = 0
            failed_count = 0
            ignore_count = 0
            row_count = 0
            import_result = {}
            err_logs = {}
            total_data = len(data)
            start_import = True
            if total_data == 0:
                start_import = False
            elif total_data <= 4:
                start_import = False
            else:
                header_type = data[0]
                header_category = data[1]
                header_key = data[2]
                if '' in header_key or '' in header_type:
                    start_import = False
            if start_import:
                last_update_sent = 0
                for index, unprocessed_row in enumerate(data):
                    no_record_flag = 0
                    msg = ""
                    if index == 0:
                        data_type = unprocessed_row
                    if index == 2:
                        key = unprocessed_row
                    if index > 3:
                        row = process_row(data_type, key, unprocessed_row)
                        row_check = False
                        req_body = {}
                        req_body["id"] = str(request_body["workflow"])
                        req_body["variables"] = {}
                        req_body["is_bulk"] = True
                        transaction_obj = Transaction.objects.create(tenant=Organisation.objects.get(id=tenant_id))
                        req_body["variables"]["transaction_id"] = str(transaction_obj.id)
                        row_count += 1
                        mandatory_field_check = True
                        for i, category in enumerate(header_category):
                            if category == 'mandatory' and row[i] == '':
                                mandatory_field_check = False
                        if all('' == column_data for column_data in row):
                            row_check = True
                        if row_check:
                            total_count += 1
                            err_logs["Data Row " + str(row_count)] = {
                                "error" : 'No data found in row no {}.'.format(row_count),
                                "message" : 'No data found in row no {}.'.format(row_count)
                            }
                            ignore_count += 1
                        elif mandatory_field_check == False:
                            total_count += 1
                            err_logs["Data Row " + str(row_count)] = {
                                "error" : 'Improper data found in row no. {}.'.format(total_count),
                                "message" : 'Improper data found in row no. {}.'.format(total_count)
                            }
                            failed_count += 1
                        else:
                            for i,j in enumerate(row):
                                if data_type[i] == 'file':
                                    file=[]
                                    if row[i]:
                                        files_list = json.loads(row[i])
                                        for doc in files_list:
                                            try:
                                                file_url = doc['url']
                                                file_name = doc['name']
                                                file_label = doc['label']
                                                res = requests.get(file_url, allow_redirects=True)
                                                content_type = res.headers['content-type']
                                                extension = '.' + file_name.split(".")[-1]
                                                file_post_url = reverse('org_forms_and_files:upload_files-list', kwargs = {'tenant' : tenant_id})
                                                file_post_url = request_body["domain_info"] + file_post_url
                                                fname =  str(uuid.uuid4()) + extension
                                                doc = SimpleUploadedFile(fname, res.content, content_type)
                                                file_obj = OrganisationFile.objects.create(
                                                    transaction_id = transaction_obj,
                                                    name = file_name,
                                                    file_label = file_label,
                                                    file = doc,
                                                    content_type=content_type,
                                                    doc_type=OrganisationFile.TYPE_CHOICES[0][0], 
                                                    tenant=tenant_id
                                                )
                                                file_post_url += '/' + str(file_obj.id)
                                                file_json = {}
                                                file_data = {}
                                                file_data['url'] = file_post_url
                                                file_data['size'] = res.headers.get('content-length', 0)
                                                file_data["form"] = ""
                                                file_data['project'] = ""
                                                file_data['baseUrl'] = "https://api.form.io"
                                                file_data['name'] = file_obj.file_label
                                                file_json['data'] = file_data
                                                file_json['name'] = file_obj.file_label
                                                file_json['originalName'] = file_obj.name
                                                file_json['size'] = res.headers.get('content-length', 0)
                                                file_json['storage'] = 'url'
                                                file_json['type'] = content_type
                                                file_json['url'] = file_post_url
                                                file.append(file_json)
                                            except Exception as error:
                                                logger.warning(str(error))
                                    req_body["variables"][key[i]] = file
                                else:
                                    req_body["variables"][key[i]] = j
                            if request_body['import_type'] == 'entity':
                                if 'enabled_for_entities_by_status' in request_body:
                                    if request_body["enabled_for_entities_by_status"] == "Active":
                                        queryset = OrganisationEntityMasterData.objects.filter(entity_model__id=request_body["master_model_id"], entity_model__tenant=tenant_id)
                                    if request_body["enabled_for_entities_by_status"] == "Inactive":
                                        queryset = OrganisationEntityMasterData.objects.deleted_set().filter(entity_model__id=request_body["master_model_id"], entity_model__tenant=tenant_id)
                                    if request_body["enabled_for_entities_by_status"] == "All":
                                        queryset = OrganisationEntityMasterData.objects.all_with_deleted().filter(entity_model__id=request_body["master_model_id"], entity_model__tenant=tenant_id)
                                else:
                                    queryset = OrganisationEntityMasterData.objects.filter(entity_model__id=request_body["master_model_id"], entity_model__tenant=tenant_id)
                                for item in request_body['unique_key']:
                                    search_key = "entity_data__" + item
                                    queryset = queryset.filter(**{ search_key : row[key.index(item)] })
                                if queryset.exists():
                                    req_body["variables"]['entity_id'] = str(queryset[0].id)
                                else:
                                    internal_error = 6090
                                    logger.error(getMessage(org_apps_errors, internal_error).format(row_count), internal_error)
                                    no_record_flag = 1
                                    msg = "No records for {} with value : {} found for Row {} ".format(request_body['unique_key'], row[key.index(item)], row_count)
                            if no_record_flag == 0:
                                if 'common_variable' in request_body and request_body['common_variable']:
                                    req_body["variables"].update(request_body['common_variable'])
                                req_body["variables"].update(request_body['context_data'])
                                url = request_body["domain_info"] + reverse('organisation_apps:app-launch-process', kwargs={'tenant': tenant_id, 'pk':request_body["workflow"]})
                                action = launch_process(req_body, tenant_id, import_obj.user.userId)
                                if action['success']:
                                    success_count += 1
                                else:
                                    err_logs["Data Row " + str(row_count)] = {
                                        "error" : action['error']['message'],
                                        "message" : action['error']['exception']
                                    }
                                    failed_count += 1
                            else:
                                err_logs["Data Row " + str(row_count)] = {
                                        "error" : msg,
                                        "message" : msg
                                    }
                                failed_count += 1
                        try:
                            percent_completed = (total_count/total_data)*100
                            if percent_completed > last_update_sent+5:
                                last_update_sent = percent_completed
                                data = {
                                    "transaction_id": str(import_obj.transaction_id),
                                    "completed": int(last_update_sent),
                                    "success": success_count,
                                    "failed": failed_count,
                                    "ignored": ignore_count
                                }
                                import_obj.result = {"completed_percentage": int(last_update_sent)}
                                import_obj.save()
                                # send_updates(tenant, import_obj.user, UpdatesConstant.BULK_PROCESS_STEP_PROGRESS, data)
                        except Exception as err:
                            internal_error = 6091
                            logger.exception(getMessage(org_apps_errors, internal_error).format(str(err)), internal_error)
                import_obj.result = {}
                import_result["success"] = success_count
                import_result["failed"] = failed_count
                import_result['ignored'] = ignore_count
                import_result["error_results"] = err_logs
                import_obj.completed_at = timezone.now()
                import_obj.result = import_result

                if failed_count:
                    import_obj.status = EntityImport.STATUS_CHOICES[2][0]
                else:
                    import_obj.status = EntityImport.STATUS_CHOICES[1][0]
                # send_updates_for_import(tenant, import_obj)
            else:
                import_obj.result = {}
                data = {
                    "error" : "Improper Excel file",
                    "message" : "Empty file found or improper header or keys"
                }
                import_result["error_results"] = {
                    'Invalid data' : data
                }
                import_obj.completed_at = timezone.now()
                import_obj.result = import_result
                import_obj.status = EntityImport.STATUS_CHOICES[2][0]
        except Exception as e:
            internal_error = 6092
            logger.warning(str(e))
            import_obj.status = EntityImport.STATUS_CHOICES[2][0]
            import_obj.completed_at = timezone.now()
            logger.exception(getMessage(org_apps_errors, internal_error).format(str(e)), internal_error)
        import_obj.save()
    else:
        internal_error = 6093
        logger.exception(getMessage(org_apps_errors, internal_error).format(entity_import_id), internal_error)
    logger.info("file_path : {0}".format(request_body["file_path"]))
    os.remove(request_body["file_path"])
    logger.info("File removed")


def parse_json(text):
    import re
    text = str(text)
    if len(text) > 0:
        text = text.strip()
        if text != "" and ((text[0] == "{" and text[-1] == "}") or (text[0] == "[" and text[-1] == "]")):
            try:
                return ",".join(json.loads(text))
            except:
                return text
        elif  bool(re.match("^[0-9]{12}@ezedox.com$",text)):
            return ""
    return text

# @app.task(bind=True, name="report_csv")
# def report_csv(self, request_body):
def report_csv(request_body, recipients, name, send_via_email):
    dirpath = tempfile.mkdtemp()
    report_path = dirpath + "/Reports.xlsx"
    workbook = xlsxwriter.Workbook(report_path)
    if isinstance(request_body, dict):
        create_report(workbook, request_body)
    elif isinstance(request_body, list):
        for body in request_body:
            create_report(workbook, body,body["sheet_name"])
    workbook.close()
    return report_path

def get_process_key_util(request, obj, tenant, transaction_id=None):
    tenant_org = Organisation.objects.get(id=tenant)
    ProcessDefinitions = process_engine.ProcessDefinitionsApi
    list_process = ProcessDefinitions.list_process_definitions
    get_process_definition = ProcessDefinitions.get_process_definition_start_form2
    req_body = {}
    req_body["latest"] = True
    req_body["key"] = obj.process_key
    action = call(module = ProcessDefinitions, func= list_process, data=req_body, tenant_id=tenant, request=request, type="get", read_replica=True)[0]
    iterate_process = action["data"]
    if iterate_process and iterate_process[0]["startFormDefined"]:
        start_form_data = {}
        start_form_data["process_definition_id"] = iterate_process[0]["id"]
        action = call(module = ProcessDefinitions, func = get_process_definition, data=start_form_data, tenant_id=tenant, request=request, type="get", read_replica=True)[0]
        response_body = {}
        response_body["formkey"] = action["key"]
        obj = request.user
        data = {}
        e_tag_id = str(uuid.uuid1())
        if transaction_id:
            try:
                transaction_obj = Transaction.objects.get(id = transaction_id, tenant=tenant)
            except Exception as error:
                internal_error = 6100
                context = {'error': str(error), 'success': False, 'message': _(getMessage(org_apps_errors, internal_error)), 'internal_error': internal_error}
                logger.error(getLogMessage(org_apps_errors, internal_error).format(error), internal_error)
                return Response(context, status=status.HTTP_400_BAD_REQUEST)
        else:
            transaction_obj = Transaction.objects.create(tenant=tenant_org)
        if obj:
            data = OrgUserSerializer(obj).data
            data["e_tag"] = e_tag_id
            data["company_name"] = tenant
            data["tenantId"] = tenant
            data.update(OrganisationDataInSerializer(tenant_org).data)
            if transaction_obj:
                data['transaction_id'] = transaction_obj.id
            response_body["data"] = data
        else:
            data["e_tag"] = e_tag_id
            if transaction_obj:
                data['transaction_id'] = transaction_obj.id
            response_body["data"] = data

        context = {"success": True, "message": _("Start Form for the workflow returned successfully."), "data": response_body}
        return Response(context, status=status.HTTP_200_OK)
    else:
        context = {"success": False, "message": _("No Start Form for the Process")}
        return Response(context, status=status.HTTP_400_BAD_REQUEST)


@app.task(bind=True, name="update_process_instance_id")
def update_process(self, transaction_id, process_id):
    try:
        transaction = Transaction.objects.get(id = transaction_id)
        transaction.process_instance_id = process_id
        transaction.save()
        if transaction_id and process_id:
            process_files = OrganisationFile.objects.filter(transaction_id = transaction_id)
            if process_files.exists():
                for file in process_files:
                    file.process_instance_id = process_id

                    new_file_key = generate_path(file,file.file.name.split('/')[-1])
                    logger.info(new_file_key)
                    storage_utils.copy_file(file, new_file_key)
                    storage_utils.delete_file(AWS_STORAGE_BUCKET_NAME, file.file.name)
                    file.file.name = new_file_key
                    file.save()
    except Exception as error:
        internal_error = 6094
        logger.exception(getMessage(org_apps_errors, internal_error).format(str(error)), internal_error)

@app.task(bind=True, name="launch_bulk_process_util")
def launch_bulk_process_util(self, request_data, tenant=None, user_email=None):
    logger.info("launch_bulk_process_util -> Inside the bulk function for data - " + json.dumps(request_data))
    variables = request_data["variables"] if "variables" in request_data else None
    domain_name = variables["domain_name"] if "domain_name" in variables else None
    action = variables["action"] if "action" in variables else None
    workflow_id_list = OrganisationWorkflow.objects.filter(kafka_topic=request_data["kafka_topic"], tenant__id=tenant)
    logger.info("launch_bulk_process_util -> Before calling the launch_process_util for {0} "
                "work flows".format(workflow_id_list.count()))
    count = 1
    for item in workflow_id_list:
        if item.kafka_domain_name is not None and item.kafka_domain_name != domain_name:
            continue
        if item.kafka_topic_action is not None and item.kafka_topic_action != action:
            continue
        logger.info("launch_bulk_process_util -> ===== Process workflow number - {0}".format(str(count)))
        logger.info("launch_bulk_process_util -> ===== Process workflow key - {0}".format(str(item.process_key)))
        try:
            request_data["id"] = item.id
            if SSL_VERIFICATION:
                response = launch_process(request_data, tenant,user_email)
            else:
                response = launch_process_util_non_http(request_data, tenant, user_email=user_email)
            logger.info("launch_bulk_process_util ->  - launch_process_util responded with => " + json.dumps(response))
        except Exception as ex:
            logger.error("launch_bulk_process_util -> Exception occurred for data - " + json.dumps(request_data))
            logger.error(traceback.format_exc())
        count += 1
    logger.info("launch_bulk_process_util -> Successfully finished processing "
                "for {0} workflows".format(str(workflow_id_list.count())))


def launch_process_util_non_http(request_data, tenant, **kwargs):
    try:
        logger.info("launch_process_util_non_http -> Inside the function for tenant => " + str(tenant))
        org = Organisation.objects.get(id=tenant)
        data = request_data
        user_email = "AnonymousUser"
        is_bulk = data.get('is_bulk', False)
        if "user_email" in kwargs:
            user_email = kwargs.get("user_email", "AnonymousUser")
        try:
            if "id" in data:
                obj = OrganisationWorkflow.objects.get(id=data.get("id", ""), tenant__id=tenant)
            else:
                obj = OrganisationWorkflow.objects.get(app_key=data.get("app_key", ""), tenant__id=tenant)
        except Exception as error:
            context = {'error': str(error), 'success': False, 'message': _('ID not found')}
            return context
        form = None
        logger.info("launch_process_util_non_http -> API call to start Process Instance")
        # API call to start Process Instance
        ProcessInstances = process_engine.ProcessInstancesApi
        create_process = ProcessInstances.create_process_instance
        req_body = dict()
        req_body.update({
            "variables": list(),
            "tenantId": tenant,
            "processDefinitionKey": obj.process_key
        })
        if obj.process_name:
            process_name = get_process_name(obj.process_name, tenant, str(obj.id), data["variables"])
            if process_name:
                req_body["name"] = process_name
            else:
                context = {'success': False, "message": _("Failed to start process"), 'error' : 'Name not formed', 'error_code' : 400}
                return context
        req_body.get("variables").append(
            {
                "name": "initiator",
                "type": "string",
                "value": user_email
            }
        )

        req_body.get("variables").append(
            {
                "name": "ezedox_api_base_url",
                "type": "string",
                "value": "{0}://{1}".format(DEFAULT_SCHEME, BASE_ORG_DOMAIN_URL)
            }
        )

        req_body.get("variables").append(
            {
                "name": "candidate_app_base_url",
                "type": "string",
                "value": "{0}://{1}".format(DEFAULT_SCHEME, CANDIDATE_DOMAIN_URL)
            }
        )

        req_body.get("variables").append(
            {
                "name": "backend_app_base_url",
                "type": "string",
                "value": BACKEND_DOMAIN_URL
            }
        )

        req_body.get("variables").append(
            {
                "name": "is_bulk",
                "type": "boolean",
                "value": is_bulk
            }
        )
        logger.info("launch_process_util_non_http -> Request Body => " + json.dumps(req_body))
        if "variables" in data:
            data["variables"]["tenantId"] = tenant
            if "transaction_id" not in data["variables"]:
                data["variables"]["transaction_id"] = str(Transaction.objects.create(tenant=org).id)
            if "e_tag" not in data["variables"]:
                data["variables"]["e_tag"] = str(uuid.uuid1())
            for var_json in data["variables"]:
                var = dict()
                var.update({
                    "name": var_json
                })

                if isinstance(data["variables"][var_json], (list, dict)):
                    var["value"] = data["variables"][var_json]
                    var["type"] = "json"
                elif form and var_json in form and form[var_json] == 'date':
                    date_str = None
                    if data["variables"][var_json]:
                        date_obj = datetime.strptime(data["variables"][var_json], '%d %b %Y')
                        date_str = date_obj.isoformat() + "Z"
                    var["value"] = date_str
                    var["type"] = "date"
                else:
                    var["value"] = data["variables"][var_json]
                req_body["variables"].append(var)
        else:
            data["variables"] = dict()
            var = dict()
            var.update({
                "name": "transaction_id",
                "type": "string",
                "value": str(Transaction.objects.create(tenant=org).id)
            })
            req_body.get("variables").append(var)
            data["variables"]['transaction_id'] = var["value"]
            var = dict()
            var.update({
                "name": "e_tag",
                "type": "string",
                "value": str(uuid.uuid1())
            })
            req_body.get("variables").append(var)
        try:
            logger.info("launch_process_util_non_http -> Hit process launch with request body")
            # logger.info(req_body)
            action = call(module=ProcessInstances, func=create_process, data=req_body, tenant_id=tenant, tenant=None,
                          request_data=request_data, type="post", user_email=user_email)
            logger.info("launch_process_util_non_http -> Action => " + str(action))
            if action[1] == 201 or action[1] == 200:
                logger.info("launch_process_util -> Inside the condition for 200 and 201")
                transaction_id = data['variables']['transaction_id']
                process_id = action[0]["id"]
                logger.info('Syncing transaction id {} with process id {}.'.format(transaction_id, process_id))
                update_process.apply_async(args=[transaction_id, process_id], priority=HIGH_PRIORITY_TASK)
                if not is_bulk:
                    pass

                if user_email != "AnonymousUser":
                    ProcessInstanceIdentityLinks = process_engine.ProcessInstanceIdentityLinksApi
                    create_process_instance_identity = ProcessInstanceIdentityLinks.create_process_instance_identity_links
                    req_body["user"] = user_email
                    req_body["type"] = "participant"
                    req_body_id = dict()
                    req_body_id["process_instance_id"] = process_id
                    try:
                        add_initiator_action = call(module=ProcessInstanceIdentityLinks,
                                                    func=create_process_instance_identity, tenant_id=tenant, id=req_body_id,
                                                    data=req_body, request_data=request_data, type="post",
                                                    user_email=user_email)
                        if add_initiator_action[1] == 201 or add_initiator_action.status_code[1] == 200:
                            logger.info(
                                "Initiator successfully added as Involved user, processinstanceId:{}".format(process_id))
                        else:
                            if action[0]['completed']:
                                logger.info(
                                    "Failed to add initiator as Involved user but current process is completed, "
                                    "processinstanceId:{}".format(process_id))
                            else:
                                internal_error = 6095
                                logger.error(getMessage(org_apps_errors, internal_error).format(process_id), internal_error)
                    except:
                        if action[0]['completed']:
                            logger.info(
                                "Failed to add initiator as Involved user but current process is completed, "
                                "processinstanceId:{}".format(process_id))
                        else:
                            internal_error = 6096
                            logger.error(getMessage(org_apps_errors, internal_error).format(process_id), internal_error)
                context = {"success": True, "message": _("Process Instance Started successfully."),
                           "data": {"id": process_id}}
                return context
            else:
                logger.info("launch_process_util -> Inside the condition for 500")
                res = action[0]
                try:
                    regex = r"(?<=No catching boundary event found for error with errorCode ').*" \
                            r"(?=', neither in same process nor in parent process)"
                    matches = re.search(regex, res["exception"])
                    if matches:
                        res["exception"] = matches.group()
                    regex2 = r'duplicate key value violates unique constraint "name_end_time_par_uniq_idx"'
                    matches2 = re.search(regex2, res["exception"])
                    if matches2:
                        res["exception"] = "Failed to start process since Ongoing Process Already present"
                    context = {'success': False, "message": _("Failed to start process"), 'error': res, 'error_code': 901}
                    return context
                except Exception as ex:
                    logger.error(ex)
                    pass
                context = {'success': False, "message": _("Failed to start process"), 'error': res}
                return context
        except Exception as error:
            context = {'success': False, "message": _("Failed to start process"), 'error': error.__str__()}
            return context
    except Exception as ex:
        logger.error(traceback.format_exc())
        context = {'success': False, "message": _("Failed to start process"), 'error': ex.__str__()}
        return context

@app.task(bind=True, name="platform_data_sync")
def platform_data_sync(self, request_data, tenant=None):
    try:
        logger.info("platform_data_sync -> Inside the function for tenant => " + str(tenant))
        data = request_data["variables"]
        kafka_topic = request_data["kafka_topic"]
        domain_name = data["domain_name"] if "domain_name" in data else None
        action = data["action"] if "action" in data else None
        process_var_sync_list = ProcessVarSync.objects.filter(
            kafka_topic__iexact=kafka_topic,
            kafka_domain_name__iexact=domain_name,
            kafka_topic_action__iexact=action
        ).filter(
            models.Q(app__tenant=tenant) | models.Q(tenant=tenant)
        )
        for item in process_var_sync_list:
            logger.info("platform_data_sync -> Inside the process_var_sync_list => " + str(item))
            workflow = [item.app] if item.app else OrganisationWorkflow.objects.filter(entity=item.entity, tenant__id=tenant)
            mapping = item.mapping
            source = item.source
            for wf in workflow:
                data_to_update = []
                process_name = get_process_name(wf.process_name, tenant, str(wf.id), data)
                process_ids = get_process_by_process_ids(process_name, tenant)
                logger.info("platform_data_sync -> process_ids: " + str(process_ids))
                for process_id in process_ids:
                    process_id = process_id["id"]
                    logger.info("platform_data_sync -> Processing process_id: " + str(process_id))
                    var_data = get_var_by_process_id(process_id)
                    source_data = data
                    if source == "GET_API_CALL":
                        employee_mgmt_service = EmployeeMgmtService(PLATFORM_BASE_URL, tenant, PLATFORM_INTERNAL_TOKEN)
                        emp_status, source_data = employee_mgmt_service.get_employee_details(data["uuid"])
                        logger.info("platform_data_sync -> response_status => " + str(emp_status))
                    for mapping_key in mapping.keys():
                        mapping_value = mapping[mapping_key]
                        source_val = None
                        if mapping_value == "MOBILE":
                            for contact in source_data["contacts"]:
                                if contact["type"] == "MOBILE":
                                    source_val = contact["contact"]
                                    break
                        elif mapping_value == "EMAIL":
                            for contact in source_data["contacts"]:
                                if contact["type"] == "EMAIL":
                                    source_val = contact["contact"]
                                    break
                        elif mapping_value.startswith("DOCUMENTS."):
                            mapping_data = mapping_value.split(".")[1]
                            mapping_data2 = mapping_value.split(".")[2]
                            for document in source_data["documents"]:
                                if document["type"] == mapping_data:
                                    source_val = document[mapping_data2]
                                    break
                        elif mapping_value.startswith("ADDRESSES."):
                            mapping_data = mapping_value.split(".")[1]
                            mapping_data2 = mapping_value.split(".")[2]
                            for address in source_data["addresses"]:
                                if address["addressType"] == mapping_data:
                                    source_val = address[mapping_data2]
                                    break
                        # elif mapping_value == "LOCATION":
                        #     defaultLocation = source_data["defaultLocation"] if "defaultLocation" in source_data else None
                        #     if defaultLocation:
                        #         cust_mgmt_service = CustomerMgmtService(PLATFORM_BASE_URL, tenant, PLATFORM_INTERNAL_TOKEN)
                        #         cust_status, cust_response = cust_mgmt_service.get_tags(defaultLocation)
                        #         logger.info("platform_data_sync -> cust_status => " + str(cust_status))
                        #         if cust_status in [200, 201]:
                        #             source_val = cust_response[0]["name"]
                        # elif mapping_value == "ROLE":
                        #     defaultRole = source_data["defaultRole"] if "defaultRole" in source_data else None
                        #     if defaultRole:
                        #         cust_mgmt_service = CustomerMgmtService(PLATFORM_BASE_URL, tenant, PLATFORM_INTERNAL_TOKEN)
                        #         cust_status, cust_response = cust_mgmt_service.get_tags(defaultRole)
                        #         logger.info("platform_data_sync -> cust_status => " + str(cust_status))
                        #         if cust_status in [200, 201]:
                        #             source_val = cust_response[0]["name"]
                        elif mapping_value == "WORK_ORDER":
                            workOrderId = source_data["workOrderId"] if "workOrderId" in source_data else None
                            if workOrderId:
                                vendor_mgmt_service = VendorMgmtService(PLATFORM_BASE_URL, tenant, PLATFORM_INTERNAL_TOKEN)
                                vendor_status, vendor_response = vendor_mgmt_service.get_workorders(workOrderId)
                                logger.info("platform_data_sync -> vendor_status => " + str(vendor_status))
                                if vendor_status in [200, 201]:
                                    source_val = vendor_response["workOrderNumber"]
                        elif mapping_value == "VENDOR_CODE":
                            vendorCode = source_data["vendorCode"] if "vendorCode" in source_data else None
                            if vendorCode:
                                vendor_mgmt_service = VendorMgmtService(PLATFORM_BASE_URL, tenant, PLATFORM_INTERNAL_TOKEN)
                                vendor_status, vendor_response = vendor_mgmt_service.get_vendor_by_vendorcode(vendorCode)
                                logger.info("platform_data_sync -> vendor_status => " + str(vendor_status))
                                if vendor_status in [200, 201]:
                                    source_val = vendor_response["vendorCode"]
                        # elif mapping_value == "WC_POLICY":
                        #     wcPolicy = source_data["wcPolicy"] if "wcPolicy" in source_data else None
                        #     if wcPolicy:
                        #         pass
                        # elif mapping_value == "LABOUR_LICENSE":
                        #     labourLicence = source_data["labourLicence"] if "labourLicence" in source_data else None
                        #     if labourLicence:
                        #         pass
                        # elif mapping_value == "REPORTING_MANAGER":
                        #     reportsTo = source_data["reportsTo"] if "reportsTo" in source_data else None
                        #     if reportsTo:
                        #         emp_status, emp_response = employee_mgmt_service.get_employee_details(reportsTo)
                        #         logger.info("platform_data_sync -> emp_status => " + str(emp_status))
                        #         if emp_status in [200, 201]:
                        #             source_val = emp_response["nameInLowerCase"]
                        elif mapping_value.startswith("EMP_NAME."):
                            mapping_data = mapping_value.split(".")[1]
                            try:
                                mapping_data2 = mapping_value.split(".")[2]
                            except:
                                mapping_data2 = "nameInLowerCase"
                            mapping_data_value = source_data[mapping_data] if mapping_data in source_data else None
                            if mapping_data_value:
                                emp_status, emp_response = employee_mgmt_service.get_employee_details(mapping_data_value)
                                logger.info("platform_data_sync -> emp_status => " + str(emp_status))
                                if emp_status in [200, 201]:
                                    source_val = emp_response[mapping_data2]
                        elif mapping_value.startswith("TAG_NAME."):
                            parts = mapping_value.split(".")
                            mapping_data = parts[1]
                            mapping_data_value = source_data[mapping_data] if mapping_data in source_data else None
                            if mapping_data_value:
                                cust_mgmt_service = CustomerMgmtService(PLATFORM_BASE_URL, tenant, PLATFORM_INTERNAL_TOKEN)
                                cust_status, cust_response = cust_mgmt_service.get_tags(mapping_data_value)
                                logger.info("platform_data_sync -> cust_status => " + str(cust_status))
                                if cust_status in [200, 201]:
                                    tag_obj = cust_response[0]
                                    if len(parts) >= 5 and parts[2] == "parents":
                                        parent_type = parts[3]
                                        parent_field = parts[4] if len(parts) > 4 else "name"
                                        parents = tag_obj.get("parents", [])
                                        for parent in parents:
                                            if parent.get("type") == parent_type:
                                                source_val = parent.get(parent_field)
                                                break
                                    else:
                                        # Case 1 & 2: TAG_NAME.<field>[.<tag_field>]
                                        tag_field = parts[2] if len(parts) > 2 else "name"
                                        source_val = tag_obj.get(tag_field)
                        else:
                            def get_nested_value(d, key_path):
                                keys = key_path.split('.')
                                for k in keys:
                                    d = d.get(k, {})
                                return d if d != {} else None
                            source_val = get_nested_value(source_data, mapping_value)
                        if source_val is not None and var_data.get(mapping_key) != source_val:
                            data_to_update.append(
                                {"name": mapping_key,
                                    "value": source_val}
                            )
                    if len(data_to_update) > 0:
                        logger.info("platform_data_sync -> process_instance_id => " + str(process_id))
                        logger.info("platform_data_sync -> data_to_update => " + str(data_to_update))
                        response_context, response_status = call(module=process_engine.ProcessInstanceVariablesApi, func=process_engine.ProcessInstanceVariablesApi.create_or_update_process_variable, tenant_id=tenant,  data={
                                                "body": data_to_update, "process_instance_id": process_id}, type="put")
                        logger.info("platform_data_sync -> process var update response_status => " + str(response_status))
    except Exception as ex:
        logger.exception("platform_data_sync -> Exception occurred: " + str(ex))

def launch_process_util(request, tenant):
    logger.info("launch_process_util -> Inside the function")
    org = Organisation.objects.get(id=tenant)
    data = request.data
    data = replace_newlines(data)
    is_bulk = data.get('is_bulk', False)
    try:
        if "id" in data:
            obj = OrganisationWorkflow.objects.get(id=data["id"], tenant__id=tenant)
        else:
            obj = OrganisationWorkflow.objects.get(app_key=data["app_key"], tenant__id=tenant)
    except Exception as error:
        context = {'error': str(error), 'success': False, 'message': _('ID not found')}
        return Response(context, status=status.HTTP_404_NOT_FOUND)

    ProcessDefinitions = process_engine.ProcessDefinitionsApi
    list_process = ProcessDefinitions.list_process_definitions
    get_process_definition = ProcessDefinitions.get_process_definition_start_form2
    req_body = {}
    req_body["latest"] = True
    req_body["key"] = obj.process_key
    action = call(module = ProcessDefinitions, func = list_process, data=req_body, tenant_id=tenant, request=request, type="get", read_replica=True)[0]
    iterate_process = action["data"]
    form = None
    if iterate_process[0]["startFormDefined"]:
        start_form_data = {}
        start_form_data["process_definition_id"] = iterate_process[0]["id"]
        action = call(module = ProcessDefinitions, func = get_process_definition, data=start_form_data, tenant_id=tenant, request= request, type="get", read_replica=True)[0]
        formkey=action["key"]
        actual_form_key, form_version = formkey.split("::")
        form=get_key_type(OrganisationForm.objects.filter(key=actual_form_key, version=form_version, tenant__id=tenant).first().keytypepair)
        if not is_bulk:
            sub_data = data["variables"]
            valid_data = ['transaction_id','e_tag','entity_id']
            transformed_data = {k: sub_data[k] for k in sub_data if k in form or k in valid_data}
            data["variables"] = transformed_data
    # API call to start Process Instance
    ProcessInstances = process_engine.ProcessInstancesApi
    create_process = ProcessInstances.create_process_instance
    req_body = {}
    req_body["variables"] = []
    if obj.process_name:
        process_name = get_process_name(obj.process_name, tenant, str(obj.id), data["variables"] if "variables" in data else [])
        if process_name:
            req_body["name"] = process_name
        else:
            context = {'success': False, "message": _("Failed to start process"), 'error' : 'Name not formed', 'error_code' : 400}
            return Response(context, status=status.HTTP_400_BAD_REQUEST)
    req_body["processDefinitionId"] = iterate_process[0]["id"]
    var = {}
    var["name"] = "initiator"
    var["type"] = "string"
    if request.user.is_anonymous:
        var["value"] = "Anonymous User"
    else:
        var["value"] = request.user.userId
    req_body["variables"].append(var)
    var1 = {}
    var1["name"] = "ezedox_api_base_url"
    var1["type"] = "string"
    var1["value"] = "{0}://{1}".format(DEFAULT_SCHEME, BASE_ORG_DOMAIN_URL)
    req_body["variables"].append(var1)
    var2 = {}
    var2["name"] = "candidate_app_base_url"
    var2["type"] = "string"
    var2["value"] = "{0}://{1}".format(DEFAULT_SCHEME, CANDIDATE_DOMAIN_URL)
    req_body["variables"].append(var2)
    var3 = {}
    var3["name"] = "backend_app_base_url"
    var3["type"] = "string"
    var3["value"] = BACKEND_DOMAIN_URL
    req_body["variables"].append(var3)
    req_body["variables"].append({"name":'is_bulk', 'type': 'boolean', 'value': is_bulk})
    if "variables" in data:
        data['variables']['tenantId'] = tenant
        if not 'transaction_id' in data["variables"]:
            data["variables"]['transaction_id'] = str(Transaction.objects.create(tenant=org).id)
        if not 'e_tag' in data['variables']:
            data['variables']['e_tag'] = str(uuid.uuid1())
        for var_json in data["variables"]:
            var = {}
            var["name"] = var_json

            if isinstance(data["variables"][var_json],(list,dict)):
                var["value"] = data["variables"][var_json]
                var["type"] =  "json"
            elif form and var_json in form and form[var_json] == 'date':
                date_str = None
                if data["variables"][var_json]:
                    date_obj = datetime.strptime(data["variables"][var_json], '%d %b %Y')
                    date_str = date_obj.isoformat() + "Z"
                var["value"] = date_str
                var["type"] =  "date"
            else:
                var["value"] = data["variables"][var_json]
            req_body["variables"].append(var)
    else:
        data["variables"] = {}
        var = {}
        var["name"] = "transaction_id"
        var["type"] = "string"
        var["value"] = str(Transaction.objects.create(tenant=org).id)
        req_body["variables"].append(var)
        data["variables"]['transaction_id'] =  var["value"]
        var = {}
        var["name"] = "e_tag"
        var["type"] = "string"
        var["value"] = str(uuid.uuid1())
        req_body["variables"].append(var)
    try:
        logger.info(req_body)
        action = call(module=ProcessInstances, func=create_process, data=req_body, tenant_id=tenant, tenant=None,
                      request=request, type="post")
        if action[1] == 201 or action[1] == 200:
            logger.info("launch_process_util -> Inside the condition for 200 and 201")
            transaction_id = data['variables']['transaction_id']
            process_id = action[0]["id"]
            logger.info('Syncing transaction id {} with process id {}.'.format(transaction_id, process_id))
            update_process.apply_async(args=[transaction_id, process_id], priority=HIGH_PRIORITY_TASK)
            if not is_bulk:
                pass
                # send_inapp_notification(request, process_id, NotificationConstant.INDIVIDUAL_TASK_CHOICE, None)
                # send_update_to_users(UpdatesConstant.UPDATE_ONGOING_PROCESS, request.tenant)

            if not request.user.is_anonymous:
                ProcessInstanceIdentityLinks = process_engine.ProcessInstanceIdentityLinksApi
                create_process_instance_identity = ProcessInstanceIdentityLinks.create_process_instance_identity_links
                req_body["user"] = request.user.email
                req_body["type"] = "participant"
                req_body_id = {}
                req_body_id["process_instance_id"] = process_id
                try:
                    add_initiator_action = call(module = ProcessInstanceIdentityLinks, func = create_process_instance_identity, tenant_id=tenant, id=req_body_id, data=req_body, request= request, type="post")
                    if add_initiator_action[1] == 201 or add_initiator_action.status_code[1] == 200:
                        logger.info("Initiator successfully added as Involved user, processinstanceId:{}".format(process_id))
                    else:
                        if action[0]['completed'] == True:
                            logger.info("Failed to add initiator as Involved user but current process is completed, processinstanceId:{}".format(process_id))
                        else:
                            internal_error = 6095
                            logger.error(getMessage(org_apps_errors, internal_error).format(process_id), internal_error)
                except:
                    if action[0]['completed'] == True:
                        logger.info("Failed to add initiator as Involved user but current process is completed, processinstanceId:{}".format(process_id))
                    else:
                        internal_error = 6096
                        logger.error(getMessage(org_apps_errors, internal_error).format(process_id), internal_error)
            else :
                logger.info("Failed to add initiator as Involved user as it was an anonymous user, processinstanceId:{}".format(process_id))
            context = {"success": True, "message": _("Process Instance Started successfully."), "data": {"id" :process_id}}
            return Response(context, status=status.HTTP_200_OK)
        else:
            logger.info("launch_process_util -> Inside the condition for 500")
            res = action[0]
            try:
                regex = r"(?<=No catching boundary event found for error with errorCode ').*(?=', neither in same process nor in parent process)"
                matches = re.search(regex, res["exception"])
                if matches:
                    res["exception"] = matches.group()
                regex2 = r'duplicate key value violates unique constraint "name_end_time_par_uniq_idx"'
                matches2 = re.search(regex2, res["exception"])
                if matches2:
                    res["exception"] = "Failed to start process since Ongoing Process Already present"
                context = {'success': False, "message": _("Failed to start process"), 'error': res, 'error_code' : 901}
                return Response(context, status=status.HTTP_400_BAD_REQUEST)
            except Exception as ex:
                logger.error(ex)
                pass
            context = {'success': False, "message": _("Failed to start process"), 'error': res}
            return Response(context, status=status.HTTP_400_BAD_REQUEST)
    except Exception as error:
        context = {'success': False, "message": _("Failed to start process"), 'error': error}
        return Response(context, status=status.HTTP_400_BAD_REQUEST)


def process_data(request, finished=None, deleted=None, tenant=None, process=None, req_json=None):

    req_body = {}
    req_body["tenantId"] = tenant
    req_body["size"] = 0
    req_body["includeProcessVariables"] = False
    if finished :
        req_body['finished'] = finished
    if deleted is not None:
        req_body["deleted"] = deleted
    if process is not None:
        req_body["processDefinitionKey"] = process
    if req_json is not None:
        for key in req_json.keys():
            req_body[key] = req_json[key]
    return req_body

def process_count(request, method, PROCESS, engine_url, finished=None, deleted=None, process=None, req_json=None):
    url = PROCESS.format(engine_url)
    req_body = {}
    req_body["tenantId"] = get_tenant(request)
    if finished :
        req_body['finished'] = finished
    if deleted is not None:
        req_body["deleted"] = deleted
    if process is not None:
        req_body["processDefinitionKey"] = process
    if req_json is not None:
        for key in req_json:
            req_body[key] = req_json[key]
    if method == "POST":
        action = requests.request("POST", url, auth=(PROCESS_ENGINE_USER, PROCESS_ENGINE_PASSWORD), data=json.dumps(req_body), headers={"Content-Type" : "application/json"})
    else:
        action = requests.get(url, auth=HTTPBasicAuth(PROCESS_ENGINE_USER, PROCESS_ENGINE_PASSWORD), params=req_body, headers={"Content-Type" : "application/json"})
    return action

def create_report(workbook, request_body,sheet_name="Sheet 1"):
    worksheet = workbook.add_worksheet(sheet_name)
    row = 0
    col = 0
    form_fields = []
    if "selected_fields" in request_body["selected_items"]:
        form_fields = request_body["selected_items"]["selected_fields"]
    process_fields = []
    if "process_fields" in request_body["selected_items"]:
        process_fields = request_body["selected_items"]["process_fields"]
    current_owner = False
    for index in range(len(process_fields)):
        if process_fields[index]['key'] == "currentOwner":
            current_owner = True
            del process_fields[index]
            break
    flag = 0
    for headers in form_fields:
        worksheet.write(row, col,headers["name"])
        col = col+1
    for headers in process_fields:
        worksheet.write(row, col,headers["name"])
        col = col+1
    worksheet.write(row, col,"PROCESS STATUS")
    col = col + 1
    if current_owner:
        worksheet.write(row, col,"Current Owner(s)")
    row = 1
    col = 0
    pids = []
    for data_instance in request_body["data"]:
        if "variables" in data_instance and len(data_instance["variables"]) == 0:
            pids.append(data_instance["id"])
            if len(pids) >= int(REPORT_BATCH_SIZE) or data_instance == request_body["data"][-1]:
                Query = process_engine.QueryApi
                query_historic_process = Query.query_historic_process_instance
                req_get_body = {}
                req_get_body["processInstanceIds"] = pids
                req_get_body["includeProcessVariables"] = True
                req_get_body["size"] = int(REPORT_BATCH_SIZE)
                req_get_body["sort"] = "startTime"
                req_get_body["order"] = "desc"
                action = call(module = Query, func= query_historic_process, data=req_get_body, tenant_id= request_body["tenantId"], type="post", read_replica=True)[0]
                for batch_process in action["data"]:
                    current_task_owner = ""
                    variable_json = {}
                    if batch_process["endTime"] is None and current_owner:
                        process_id = batch_process["id"]
                        Tasks_api = process_engine.TasksApi
                        list_task = Tasks_api.list_tasks
                        query_params = {}
                        query_params['process_instance_id'] = process_id
                        res_action = call(module = Tasks_api, func= list_task, data=query_params, tenant_id= request_body["tenantId"], type="get", read_replica=True)[0]
                        tasks = res_action["data"]
                        index = 0
                        for task in tasks:
                            assignee_value =""
                            if not task['assignee']:
                                task_id = {}
                                task_id["task_id"] = task['id']
                                TaskIdentityLinks = process_engine.TaskIdentityLinksApi
                                list_tasks_instance_identity = TaskIdentityLinks.list_tasks_instance_identity_links
                                group_data = call(module = TaskIdentityLinks, func= list_tasks_instance_identity, data=task_id, tenant_id= request_body["tenantId"], type="get", read_replica=True)[0]
                                group_id = group_data[0]["group"]
                                group = OrganisationGroup.objects.get(id = group_id)
                                if index == 0:
                                    assignee_value = group.name + "(" + task["name"] + ")"
                                    index+=1
                                else:
                                    assignee_value = ","+ group.name + "(" + task["name"] + ")"
                            else:
                                if index == 0:
                                    assignee_value = task['assignee'] + "(" + task["name"] + ")"
                                    index+=1
                                else:
                                    assignee_value = ","+ task['assignee'] + "(" + task["name"] + ")"
                            current_task_owner += assignee_value


                    for variable_values in batch_process["variables"]:
                        if 'type' in variable_values and variable_values["type"] == 'date':
                            date_isoformat = variable_values['value']
                            date_obj = datetime.strptime(date_isoformat,"%Y-%m-%dT%H:%M:%SZ")
                            date_str = date_obj.strftime('%d %b %Y')
                            variable_json[variable_values["name"]] = date_str
                        else:
                            variable_json[variable_values["name"]] = variable_values['value']
                    for headers in form_fields:
                        if headers["key"] in variable_json:
                            flag = 1
                            if type(variable_json[headers["key"]]) == bool:
                                if variable_json[headers["key"]] == True :
                                    variable_json[headers["key"]] = "TRUE"
                                else :
                                    variable_json[headers["key"]] = "FALSE"
                            worksheet.write(row, col,parse_json(variable_json[headers["key"]]))
                        col = col + 1
                    for headers in process_fields:
                        if headers["key"] in batch_process:
                            flag = 1
                            if headers["key"] == "startTime" or headers["key"] =="endTime":
                                if batch_process[headers["key"]]:
                                    date_str = batch_process[headers["key"]]
                                    worksheet.write(row, col,date_str)
                                else:
                                    worksheet.write(row, col,batch_process[headers["key"]])
                            else:
                                worksheet.write(row, col,batch_process[headers["key"]])
                        col = col + 1
                    if batch_process["endTime"] is None:
                        worksheet.write(row, col,"ONGOING")
                    else:
                        if batch_process["deleteReason"] is not None:
                            worksheet.write(row, col,"WITHDRAWN")
                        else:
                            worksheet.write(row, col,"COMPLETED")
                    col=col+1
                    if current_owner:
                        worksheet.write(row, col,current_task_owner)
                        current_task_owner =""
                    if flag == 1 :
                        row = row + 1
                        col = 0
                        flag = 0
                pids = []
        # else:
        #     variable_json = {}

        #     for variable_values in data_instance["variables"]:
        #         variable_json[variable_values["name"]] = variable_values['value']
        #     for headers in form_fields:
        #         if headers["key"] in variable_json:
        #             flag = 1
        #             if type(variable_json[headers["key"]]) == bool:
        #                 if variable_json[headers["key"]] == True :
        #                     variable_json[headers["key"]] = "TRUE"
        #                 else :
        #                     variable_json[headers["key"]] = "FALSE"
        #             worksheet.write(row, col,parse_json(variable_json[headers["key"]]))
        #         col = col + 1
        #     for headers in process_fields:
        #         if headers["key"] in data_instance:
        #             flag = 1
        #             worksheet.write(row, col,data_instance[headers["key"]])
        #         col = col + 1
        #     if data_instance["endTime"] is None:
        #         worksheet.write(row, col,"ONGOING")
        #     else:
        #         if data_instance["deleteReason"] is not None:
        #             worksheet.write(row, col,"WITHDRAWN")
        #         else:
        #             worksheet.write(row, col,"COMPLETED")
        if flag == 1:
            row = row + 1
            col = 0
            flag = 0


def get_system_filter_value(filter_field,user_data):
    try:
        exect_value = []
        if filter_field =='entity_location':
            user_location_data = user_data[0].location
            if user_location_data:
                exect_value.append(user_location_data.name)
        elif filter_field =='entity_department':
            user_department_data = user_data[0].department
            if user_department_data:
                exect_value.append(user_department_data.name)
        else:
            user_extra_field_data  = user_data[0].extra_fields
            if user_extra_field_data:
                filter_field_value = user_extra_field_data[filter_field]
                if filter_field_value:
                    if type(filter_field_value) == dict and filter_field_value['value']:
                        exect_value.append(filter_field_value['value'])
                    elif type(filter_field_value) == list:
                        for list_data in filter_field_value:
                            if list_data['value']:
                                exect_value.append(list_data['value'])
                    else:
                        exect_value.append(filter_field_value)
        return exect_value
    except Exception as error:
        internal_error = 6097
        logger.error(getMessage(org_apps_errors, internal_error).format(error), internal_error)

def get_filter_task_title(taskTitle):
    try:
        if taskTitle is not None:
            task_title_list = re.split(r'(\s+)', taskTitle)
            filter_task_title = []
            for i in task_title_list:
                if i.startswith("${") and i.endswith("}"):
                    i = "%"
                filter_task_title.append(i)
            filter_task_title = "".join(filter_task_title)
            return filter_task_title
        else:
            return None
    except Exception as error:
        internal_error = 6098
        logger.error(getMessage(org_apps_errors, internal_error).format(error), internal_error)



def utc_date_conversion(date_str):
    utc_zone = tz.gettz('UTC')
    TIME_ZONE = tz.gettz('Asia/Kolkata')
    date_str = datetime.strptime(date_str, '%Y-%m-%d').strftime("%Y-%m-%d %H:%M:%S")
    local_time = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
    local_time = local_time.replace(tzinfo=TIME_ZONE)
    utc_time = local_time.astimezone(utc_zone)
    utc_date_string = utc_time.strftime("%Y-%m-%dT%H:%M:%SZ")
    return utc_date_string

def ist_date_conversion(date_str):
    utc_zone = tz.gettz('UTC')
    TIME_ZONE = tz.gettz('Asia/Kolkata')
    date_str = datetime.strptime(date_str,  "%Y-%m-%dT%H:%M:%S.%fZ").strftime("%Y-%m-%d %H:%M:%S")
    local_time = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
    local_time = local_time.replace(tzinfo=utc_zone)
    ist_time = local_time.astimezone(TIME_ZONE)
    ist_date_string = ist_time.strftime("%Y-%m-%d %H:%M:%S")
    return ist_date_string

def launch_process(request_data, tenant, initiator):
    try:
        logger.info("Inside Launch Process")
        org = Organisation.objects.get(id=tenant)
        data = request_data
        is_bulk = data.get('is_bulk', False)
        if "id" in data:
            obj = OrganisationWorkflow.objects.get(id=data["id"], tenant__id=tenant)
        else:
            obj = OrganisationWorkflow.objects.get(app_key=data["app_key"], tenant__id=tenant)
        # API call to start Process Instance
        ProcessInstances = process_engine.ProcessInstancesApi
        create_process = ProcessInstances.create_process_instance
        req_body = {}
        req_body["tenantId"] = tenant
        if obj.process_name:
            process_name = get_process_name(obj.process_name, tenant, str(obj.id), data["variables"])
            if process_name:
                req_body["name"] = process_name
            else:
                context = {'success': False, "message": _("Failed to start process"), 'error' : 'Name not formed', 'error_code' : 400}
                return context
        req_body["variables"] = []
        req_body["processDefinitionKey"] = obj.process_key
        var = {}
        var["name"] = "initiator"
        var["type"] = "string"
        var["value"] = initiator
        req_body["variables"].append(var)
        var1 = {}
        var1["name"] = "ezedox_api_base_url"
        var1["type"] = "string"
        var1["value"] = "{0}://{1}".format(DEFAULT_SCHEME, BASE_ORG_DOMAIN_URL)
        req_body["variables"].append(var1)
        var2 = {}
        var2["name"] = "candidate_app_base_url"
        var2["type"] = "string"
        var2["value"] = "{0}://{1}".format(DEFAULT_SCHEME, CANDIDATE_DOMAIN_URL)
        req_body["variables"].append(var2)
        var3 = {}
        var3["name"] = "backend_app_base_url"
        var3["type"] = "string"
        var3["value"] = BACKEND_DOMAIN_URL
        req_body["variables"].append(var3)
        req_body["variables"].append({"name":'is_bulk', 'type': 'boolean', 'value': is_bulk})
        if "variables" in data:
            data['variables']['tenantId'] = tenant
            if not 'transaction_id' in data["variables"]:
                data["variables"]['transaction_id'] = str(Transaction.objects.create(tenant=org).id)
            if not 'e_tag' in data['variables']:
                data['variables']['e_tag'] = str(uuid.uuid1())
            for var_json in data["variables"]:
                var = {}
                var["name"] = var_json
                if isinstance(data["variables"][var_json],(list,dict)):
                    var["value"] = data["variables"][var_json]
                    var["type"] =  "json"
                else:
                    var["value"] = data["variables"][var_json]
                req_body["variables"].append(var)
        else:
            data["variables"] = {}
            var = {}
            var["name"] = "transaction_id"
            var["type"] = "string"
            var["value"] = str(Transaction.objects.create(tenant=org).id)
            req_body["variables"].append(var)
            data["variables"]['transaction_id'] =  var["value"]
            var = {}
            var["name"] = "e_tag"
            var["type"] = "string"
            var["value"] = str(uuid.uuid1())
            req_body["variables"].append(var)

        logger.info(req_body)
        action = call(module = ProcessInstances, func = create_process, data=req_body, tenant_id=tenant, tenant=None, type="post")
        if action[1] == 201 or action[1] == 200:
            logger.info("Success Api call")
            transaction_id = data['variables']['transaction_id']
            process_id = action[0]["id"]
            logger.info('Syncing transaction id {} with process id {}.'.format(transaction_id, process_id))
            update_process.apply_async(args=[transaction_id, process_id], priority=HIGH_PRIORITY_TASK)
            context = {"success": True, "message": _("Process Instance Started successfully."), "data": {"id" :process_id}}
            return context
        else:
            logger.info("Failed Api call")
            res = action[0]
            try:
                regex = r"(?<=No catching boundary event found for error with errorCode ').*(?=', neither in same process nor in parent process)"
                matches = re.search(regex, res["exception"])
                if matches:
                    res["exception"] = matches.group()
                context = {'success': False, "message": _("Failed to start process"), 'error': res, 'error_code' : 901}
                return context
            except:
                pass
            context = {'success': False, "message": _("Failed to start process"), 'error': res}
            return context
        
    except Exception as error:
        context = {'success': False, "message": _("Failed to start process"), 'error': error}
        logger.info(context)
        return context

@app.task(bind=True, name="bulk_task_complete")
def bulk_task_complete(self, task_id_list, request_body_complete, request_body_claim, tenant_id):
    request_body_unclaim = {"action": "claim"}
    for task_id in task_id_list:
        try:
            response_claim, response_claim_status_code  = call(module=process_engine.TasksApi, func=process_engine.TasksApi.execute_task_action, data={"task_id": task_id, "body": request_body_claim}, tenant_id=tenant_id, type="put")
            response_complete, response_complete_status_code = call(module=process_engine.TasksApi, func=process_engine.TasksApi.execute_task_action, data={"task_id": task_id, "body": request_body_complete}, tenant_id=tenant_id, type="put")
            if response_complete_status_code >= 300:
                response_unclaim, response_unclaim_status_code  = call(module=process_engine.TasksApi, func=process_engine.TasksApi.execute_task_action, data={"task_id": task_id, "body": request_body_unclaim}, tenant_id=tenant_id, type="put")
        except Exception as e:
            logger.exception(
                "Unexpected error occurred while calling task complete api - {0}".format(e))

def hash_string(s):
    # Simple hash function using SHA-256
    sha256 = hashlib.sha256()
    sha256.update(s.encode('utf-8'))
    return int(sha256.hexdigest(), 16)


def determine_partition(uuid, num_partitions):
    # Hash the UUID and take the modulo to determine the partition
    hash_value = hash_string(str(uuid))
    partition_number = hash_value % num_partitions
    return partition_number
